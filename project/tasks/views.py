from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView, CreateView, ListView, UpdateView, DetailView, DeleteView
from .models import *
from .forms import TaskForm
import logging
logger=logging.getLogger("tasks")

class HomeView(TemplateView):
    template_name = 'home.html'


class SignUpView(CreateView):
    form_class = UserCreationForm
    template_name = 'registration/signup.html'
    success_url = reverse_lazy('tasks:list')

    def form_valid(self, form):
        response = super().form_valid(form)
        user=self.object
        login(self.request, user)
        messages.success(self.request, 'Welcome!! Thank you for signing up!')
        return response


class OwnerQuerysetMixin(LoginRequiredMixin):
    def get_queryset(self):
        qs=super().get_queryset()
        return qs.filter(owner=self.request.user)

class OwnerObjectMixin(UserPassesTestMixin):
    def test_func(self):
        obj=self.get_object()
        return obj.owner_id == self.request.user.id

    def handle_no_permission(self):
        messages.error(self.request, 'Sorry, you are not authorized to perform this action.')
        return super().handle_no_permission()



class TaskListView(OwnerQuerysetMixin,ListView):
    model = Task
    template_name = 'tasks/list.html'
    context_object_name = 'tasks'
    paginate_by = 4

    def get_queryset(self):
        base=super().get_queryset()
        return filtered_tasks_qs(self.request,base)



class TaskCreateView(LoginRequiredMixin,CreateView):
    model = Task
    form_class = TaskForm
    template_name = "tasks/form.html"
    success_url = reverse_lazy('tasks:list')

    def form_valid(self, form):
        form.instance.owner = self.request.user
        messages.success(self.request, 'Task created!')
        return super().form_valid(form)
    
    
class TaskUpdateView(OwnerQuerysetMixin,OwnerObjectMixin,UpdateView):
    model = Task
    form_class = TaskForm
    template_name = "tasks/form.html"
    success_url = reverse_lazy('tasks:list')
    def form_valid(self, form):
        messages.success(self.request, 'Task updated!')
        return super().form_valid(form)



class TaskDetailView(OwnerQuerysetMixin,OwnerObjectMixin,DetailView):
    model = Task
    template_name = "tasks/detail.html"
    context_object_name = 'task'

class TaskDeleteView(OwnerQuerysetMixin,OwnerObjectMixin,DeleteView):
    model = Task
    template_name = "tasks/confirm_delete.html"
    success_url = reverse_lazy('tasks:list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Task deleted!')
        return super().delete(request, *args, **kwargs)



class TaskCompleteView(LoginRequiredMixin,View):
    @method_decorator(require_POST)
    def post(self,request,pk: int):
        task=get_object_or_404(Task, pk=pk,owner=self.request.user)
        if task.status == Task.Status.DONE:
            messages.info(self.request, 'Task have already done')
            return redirect('tasks:list')

        task.status = Task.Status.DONE
        if not task.completed_at:
            task.completed_at = timezone.now()
        task.save(update_fields=['status','completed_at','updated_at'])
        logger.info('Task completed successfully id=%s owner_id=%s title=%s', task.pk, self.request.user.id, task.title)
        messages.success(self.request, 'Task completed!')
        return redirect('tasks:list')




def filtered_tasks_qs(request,base_qs):
    qs=base_qs.select_related('owner').prefetch_related("tags")
    q=request.GET.get('q','').strip()
    if q:
        qs=qs.filter(title__icontains=q) | qs.filter(description__icontains=q)
    status=request.GET.get('status')
    valid_statuses={c[0] for c in Task.Status.choices}
    if status in valid_statuses:
        qs=qs.filter(status=status)

    priority = request.GET.get('priority')
    if priority and priority.isdigit() and int(priority) in {1,2,3}:
        qs=qs.filter(priority=int(priority))

    qs=qs.order_by('-created_at')
    return qs






from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class TaskExportExcelView(LoginRequiredMixin,View):
    def get(self,request):
        base=Task.objects.filter(owner=request.user)
        qs=filtered_tasks_qs(request,base)

        wb = Workbook()
        ws=wb.active
        ws.title = 'Tasks'
        headers=["ID","Title","Status","Priority","Due date","Completed at","Tags","Created","Updated"]
        ws.append(headers)
        for t in qs:
            tags_str=", ".join(t.tags.values_list("name",flat=True))
            ws.append([
                t.id,
                t.title,
                t.get_status_display(),
                t.get_priority_display(),
                t.due_date.isoformat() if t.due_date else "",
                t.completed_at.isoformat(sep=" ") if t.completed_at else "",
                tags_str,
                t.created_at.strftime("%m/%d/%Y %H:%M"),
                t.updated_at.strftime("%m/%d/%Y %H:%M"),
            ])
        for col in range(1,len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        resp=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"]='attachment; filename="tasks.xlsx"'
        wb.save(resp)
        return resp
